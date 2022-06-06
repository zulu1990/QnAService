from typing import Tuple
from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet.cell_range import CellRange



COUPA: str = "coupa_"
CENTRAL_BUYING = "central_buying_"
CONCUR = "concur_"
CITICARD = "citicard_"
FINANCE = "finance_"

is_coupa = lambda value: value.startswith(COUPA)
is_central_buying = lambda value: value.startswith(CENTRAL_BUYING)
is_concur = lambda value: value.startswith(CONCUR)
is_citicard = lambda value: value.startswith(CITICARD)
is_finance = lambda value: value.startswith(FINANCE)

def get_type(value):
    if is_coupa(value): return 'coupa'
    if is_central_buying(value): return 'coupa'
    if is_concur(value): return 'concur'
    if is_citicard(value): return 'citicard'
    if is_finance(value): return 'finance'

def is_training_line(value: str) -> bool:
    return is_coupa(value) \
        or is_central_buying(value) \
        or is_concur(value) \
        or is_citicard(value) \
        or is_finance(value)

def is_in_range(cell: Cell, range: CellRange) -> bool:
    return cell.column >= range.min_col and \
        cell.column <= range.max_col  and \
        cell.row >= range.min_row and \
        cell.row <= range.max_row


def get_cell_value(cell: Cell, sheet) -> str:
    if type(cell).__name__ != MergedCell.__name__:
        return cell.value

    merged_ranges = sheet.merged_cells.ranges
    target_range = next((range for range in merged_ranges if is_in_range(cell, range)), None)

    if target_range is None:
        raise ValueError("Merged cell wasn't found in merged cells ranges")


    coords = [x for x in target_range.cells]
    first_cell = sheet[coords[0][0]][coords[0][1] - 1]

    if first_cell.value is None:
        return cell.value

    return first_cell is None if cell.value else first_cell.value


def split_training_string(value):
    if not isinstance(value, str):
        return []
    
    return value.split('\n')


# in case if someone forgets remove extra line
# causes error in qna maker
def split_and_remove_empty_string(value, meta):
    split_string = split_training_string(value)
    if len(split_string) > 0 and [x for x in split_string if x == '']:
        print(f"Empty Line Detected in qna_maker cell on {meta}")
        raise Exception(f"Empty Line Detected in qna_maker cell on {meta}")
    return split_string


def split_meta_string(meta_list: list):
    return {item.split(':')[0] : item.split(':')[1] for item in meta_list}

def validate_meta_info(meta):
    if "id" not in meta or "country" not in meta:
        raise Exception('Metatag missing information on id or country')
    return

def extract_sheet_data(work_sheet) -> Tuple:
    all_data = []

    for cell in work_sheet["A"]:
        if (cell.value and is_training_line(cell.value)):
            row = work_sheet[cell.row]
            meta = split_meta_string(split_training_string(row[1].value))
            meta['id'] = row[0].value
            validate_meta_info(meta)
            training_data = {}
            training_data['name'] = row[0].value
            training_data['type'] = get_type(row[0].value)
            training_data['meta'] = meta
            training_data['luis'] = split_training_string(get_cell_value(row[2], work_sheet))
            training_data['qna_maker'] = split_and_remove_empty_string(get_cell_value(row[3], work_sheet), row[0].value)
            training_data['verification'] = split_training_string(get_cell_value(row[4], work_sheet))
            training_data['answer'] = row[5].value

            all_data.append(training_data)

    return all_data

def parse_excel_training(path: str, sheet_name: str):
    wb = load_workbook(path)
    work_sheet = wb[sheet_name]

    return extract_sheet_data(work_sheet=work_sheet)